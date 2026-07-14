#!/usr/bin/env python3
"""
みっちー財政カルテ 年次データ更新スクリプト

使い方:
  1. scripts/config.json のURLを新年度のものに書き換える
  2. GitHub Actionsの「Run workflow」ボタンを押す（または python scripts/update_fiscal_data.py を手動実行）

このスクリプトは:
  - config.json のURLから総務省Excelをダウンロード
  - 財政指標・財政調整基金・人口・歳出歳入・目的別歳出・年齢階級別人口を解析
  - data-*.json 6ファイルの値を更新（履歴は自動でスライド、8年分までは追加、それ以降はローリング）
  - 更新後に整合性チェックを行い、異常があれば処理を中断してエラー終了する（コミットさせない）
"""

import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "scripts" / "config.json"
DOWNLOAD_DIR = ROOT / "scripts" / "_downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

REGION_FILES = {
    "data-hokkaido-tohoku.json": ["北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"],
    "data-kanto.json": ["茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県"],
    "data-chubu.json": ["新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県", "静岡県", "愛知県"],
    "data-kinki.json": ["三重県", "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県"],
    "data-chugoku-shikoku.json": ["鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県"],
    "data-kyushu.json": ["福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"],
}


# ---------- 共通ユーティリティ ----------

def download(url, name):
    path = DOWNLOAD_DIR / f"{name}.xlsx"
    print(f"  ダウンロード中: {name} <- {url}")
    urllib.request.urlretrieve(url, path)
    return path


def normalize(name):
    return re.sub(r"（.*?）", "", str(name))


def lookup_muni(idx, pref, name):
    cand = normalize(name)
    for variant in (cand, cand.replace("ヶ", "ケ"), cand.replace("ケ", "ヶ")):
        if (pref, variant) in idx:
            return idx[(pref, variant)]
    known_variant = KNOWN_NAME_VARIANTS.get((pref, cand))
    if known_variant and (pref, known_variant) in idx:
        return idx[(pref, known_variant)]
    return None


def strip_district(name):
    """郡名プレフィックスを除去（例: 石狩郡当別町 -> 当別町）"""
    return re.sub(r"^.*?郡", "", str(name))


# 北方領土の「幽霊自治体」（国後郡泊村など）は実在の同名自治体と衝突するため、
# 郡名除去の対象から除外する（例: 古宇郡泊村＝実在 と 国後郡泊村＝北方領土 の衝突回避）
NORTHERN_TERRITORY_DISTRICTS = ("国後郡", "択捉郡", "色丹郡", "紗那郡", "蕊取郡")

# 異体字・表記ゆれで自動マッチしない既知の自治体（正式表記 -> 総務省Excel側の表記）
KNOWN_NAME_VARIANTS = {
    ("高知県", "梼原町"): "檮原町",
    ("福岡県", "須恵町"): "須惠町",
}


def is_northern_territory_row(name):
    return any(str(name).startswith(d) for d in NORTHERN_TERRITORY_DISTRICTS)


def slide_and_set(entry, prefix, new_value, start_idx=1, max_total=8):
    """履歴フィールドをスライドし、新しい主値をセットする。
    8データポイント（履歴7+最新1）に達するまでは追加、以降は最古を破棄してローリング。
    ※ new_value が現在の主値と完全に同じ場合は「新年度データではない」と判断し、
      スライドせずスキップする（同じ総務省データで誤って再実行した際の重複防止）。
    """
    old_main = entry.get(prefix)
    if new_value is not None and old_main is not None and new_value == old_main:
        return  # 値が変わっていない＝重複実行の可能性が高いのでスライドしない
    n = 0
    while f"{prefix}_r{start_idx + n}" in entry:
        n += 1
    if n < max_total - 1:
        if old_main is not None:
            entry[f"{prefix}_r{start_idx + n}"] = old_main
    else:
        for i in range(start_idx, start_idx + n - 1):
            entry[f"{prefix}_r{i}"] = entry.get(f"{prefix}_r{i + 1}")
        if old_main is not None:
            entry[f"{prefix}_r{start_idx + n - 1}"] = old_main
    if new_value is not None:
        entry[prefix] = new_value


# ---------- Excel解析 ----------

def parse_fiscal_indicators(muni_path, pref_path):
    muni_idx, pref_idx = {}, {}
    wb = openpyxl.load_workbook(muni_path, data_only=True)
    for row in wb.active.iter_rows(min_row=3, values_only=True):
        code, pref, name, f, x, d, u = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if not name:
            continue
        muni_idx[(pref, name)] = {"f": f, "x": x, "d": d, "u": u}
    wb = openpyxl.load_workbook(pref_path, data_only=True)
    for row in wb.active.iter_rows(min_row=3, values_only=True):
        pref, f, x, d, u = row[0], row[1], row[2], row[3], row[4]
        if not pref:
            continue
        pref_idx[pref] = {"f": f, "x": x, "d": d, "u": u}
    return muni_idx, pref_idx


def parse_reserve_fund(muni_path, pref_path):
    muni_idx, pref_idx = {}, {}
    wb = openpyxl.load_workbook(muni_path, data_only=True)
    for row in wb.active.iter_rows(min_row=6, values_only=True):
        code, pref, name = row[1], row[2], row[3]
        if not name:
            continue
        muni_idx[(pref, name)] = row[8]  # 千円 -> 後で /100000 して億円に
    wb = openpyxl.load_workbook(pref_path, data_only=True)
    for row in wb.active.iter_rows(min_row=6, values_only=True):
        pref, name, val = row[4], row[5], row[10]
        if not pref or name not in ("－", "-"):
            continue
        pref_idx[pref] = val
    return muni_idx, pref_idx


def parse_population(muni_path):
    """人口ファイルは市町村・都道府県データが1本にまとまっているため、ファイルは1つでよい。"""
    muni_idx, pref_idx = {}, {}
    wb = openpyxl.load_workbook(muni_path, data_only=True)
    for row in wb.active.iter_rows(min_row=7, values_only=True):
        pref, name = row[1], row[2]
        if not pref or pref == "合計":
            continue
        pop, growth = row[5], row[20]
        entry = {"pop": pop, "g": round(growth, 3) if growth is not None else None}
        if name == "-":
            pref_idx[pref] = entry
        else:
            if is_northern_territory_row(name):
                continue  # 北方領土の幽霊自治体は実在自治体名と衝突するため除外
            stripped = strip_district(name)
            muni_idx[(pref, stripped)] = entry
            muni_idx[(pref, name)] = entry
    return muni_idx, pref_idx


def parse_budget(city_path, town_path, pref_path):
    muni_idx = {}
    for path in (city_path, town_path):
        wb = openpyxl.load_workbook(path, data_only=True)
        cur_pref = None
        for row in wb.active.iter_rows(min_row=16, values_only=True):
            code, name = row[14], row[15]
            if not name:
                continue
            name = str(name).replace("\u3000", "").strip()
            if not code or str(code).strip() == "":
                cur_pref = name
                continue
            ei, eo = row[39], row[40]  # 列40=歳入総額, 列41=歳出総額（2026-07修正: 従来逆だった）
            if eo is None:
                continue
            muni_idx[(cur_pref, name)] = (eo, ei)
    pref_idx = {}
    wb = openpyxl.load_workbook(pref_path, data_only=True)
    for row in wb.active.iter_rows(min_row=7, values_only=True):
        pref = row[0]
        if not pref:
            continue
        pref = str(pref).replace("\u3000", "").strip()
        pref_idx[pref] = (row[2], row[1])  # 2026-07修正: 市町村と同じ列順の疑いがあり入れ替え（要検証）
    return muni_idx, pref_idx


def parse_purpose_expenditure(city_path, town_path, pref_path):
    muni_idx = {}
    for path in (city_path, town_path):
        wb = openpyxl.load_workbook(path, data_only=True)
        cur_pref = None
        for row in wb.active.iter_rows(min_row=16, values_only=True):
            code, name = row[14], row[15]
            if not name:
                continue
            name = str(name).replace("\u3000", "").strip()
            if not code or str(code).strip() == "":
                cur_pref = name
                continue
            minsei, jido, edu = row[24], row[27], row[57]
            if edu is None:
                continue
            muni_idx[(cur_pref, name)] = (minsei, jido, edu)
    pref_idx = {}
    wb = openpyxl.load_workbook(pref_path, data_only=True)
    for row in wb.active.iter_rows(min_row=6, values_only=True):
        pref = row[0]
        if not pref:
            continue
        pref = str(pref).replace("\u3000", "").strip()
        pref_idx[pref] = (row[13], row[16], row[56], row[1])  # minsei, jido, edu, eo
    return muni_idx, pref_idx


def parse_age_population(muni_path, pref_path):
    muni_idx = {}
    wb = openpyxl.load_workbook(muni_path, data_only=True)
    for row in wb.active.iter_rows(min_row=7, values_only=True):
        pref, name, sex = row[1], row[2], row[3]
        if sex != "計" or not name or name == "-":
            continue
        p0, p5, p10, p15 = row[5], row[6], row[7], row[8]
        if p0 is None:
            continue
        if is_northern_territory_row(name):
            continue  # 北方領土の幽霊自治体は実在自治体名と衝突するため除外
        under18 = p0 + p5 + p10 + p15 * 3 / 5
        stripped = strip_district(name)
        muni_idx[(pref, stripped)] = under18
        muni_idx[(pref, str(name))] = under18
    pref_idx = {}
    wb = openpyxl.load_workbook(pref_path, data_only=True)
    for row in wb.active.iter_rows(min_row=7, values_only=True):
        pref, sex = row[1], row[2]
        if sex != "計" or not pref:
            continue
        p0, p5, p10, p15 = row[4], row[5], row[6], row[7]
        pref_idx[pref] = p0 + p5 + p10 + p15 * 3 / 5
    return muni_idx, pref_idx


# ---------- メイン処理 ----------

def main():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))

    print("① Excelファイルをダウンロード中...")
    fi_muni_p = download(config["fiscal_indicators"]["muni"], "fi_muni")
    fi_pref_p = download(config["fiscal_indicators"]["pref"], "fi_pref")
    rf_muni_p = download(config["reserve_fund"]["muni"], "rf_muni")
    rf_pref_p = download(config["reserve_fund"]["pref"], "rf_pref")
    pop_muni_p = download(config["population"]["muni"], "pop_muni")
    bg_city_p = download(config["budget"]["city"], "bg_city")
    bg_town_p = download(config["budget"]["town"], "bg_town")
    bg_pref_p = download(config["budget"]["pref"], "bg_pref")
    pe_city_p = download(config["purpose_expenditure"]["city"], "pe_city")
    pe_town_p = download(config["purpose_expenditure"]["town"], "pe_town")
    pe_pref_p = download(config["purpose_expenditure"]["pref"], "pe_pref")
    ap_muni_p = download(config["age_population"]["muni"], "ap_muni")
    ap_pref_p = download(config["age_population"]["pref"], "ap_pref")

    print("② Excelを解析中...")
    fi_muni, fi_pref = parse_fiscal_indicators(fi_muni_p, fi_pref_p)
    rf_muni, rf_pref = parse_reserve_fund(rf_muni_p, rf_pref_p)
    pop_muni, pop_pref = parse_population(pop_muni_p)
    bg_muni, bg_pref = parse_budget(bg_city_p, bg_town_p, bg_pref_p)
    pe_muni, pe_pref = parse_purpose_expenditure(pe_city_p, pe_town_p, pe_pref_p)
    ap_muni, ap_pref = parse_age_population(ap_muni_p, ap_pref_p)

    print("③ data-*.json を更新中...")
    warnings = []
    for fname, pref_keys in REGION_FILES.items():
        path = ROOT / fname
        db = json.loads(path.read_text(encoding="utf-8"))
        for name, entry in db.items():
            is_pref = name in pref_keys
            pref = entry.get("p")

            fi = fi_pref.get(name) if is_pref else lookup_muni(fi_muni, pref, name)
            rf = rf_pref.get(name) if is_pref else lookup_muni(rf_muni, pref, name)
            pp = pop_pref.get(name) if is_pref else lookup_muni(pop_muni, pref, name)
            bg = bg_pref.get(name) if is_pref else lookup_muni(bg_muni, pref, name)
            pe = pe_pref.get(name) if is_pref else lookup_muni(pe_muni, pref, name)
            ap = ap_pref.get(name) if is_pref else lookup_muni(ap_muni, pref, name)

            if fi:
                slide_and_set(entry, "f", fi["f"], start_idx=1)
                slide_and_set(entry, "x", fi["x"], start_idx=1)
                slide_and_set(entry, "d", fi["d"], start_idx=1)
                slide_and_set(entry, "u", fi["u"] if fi["u"] not in (None, "-", "－") else None, start_idx=1)
            else:
                warnings.append(f"{fname}:{name} 財政指標が見つかりません")

            if rf is not None:
                slide_and_set(entry, "r", round(rf / 100, 1), start_idx=2)
            else:
                warnings.append(f"{fname}:{name} 財政調整基金が見つかりません")

            if pp:
                slide_and_set(entry, "pop", pp["pop"], start_idx=2)
                slide_and_set(entry, "g", pp["g"], start_idx=2)
            else:
                warnings.append(f"{fname}:{name} 人口が見つかりません")

            eo_oku = ei_oku = None
            if bg:
                eo_oku, ei_oku = round(bg[0] / 100000, 1), round(bg[1] / 100000, 1)
                slide_and_set(entry, "eo", eo_oku, start_idx=1)
                slide_and_set(entry, "ei", ei_oku, start_idx=1)
            else:
                warnings.append(f"{fname}:{name} 歳出歳入が見つかりません")

            if pe and eo_oku:
                minsei, jido, edu_exp = pe[0], pe[1], pe[2]
                edu_ratio = round(edu_exp / (eo_oku * 100000) * 100, 1) if eo_oku else None
                slide_and_set(entry, "edu", edu_ratio, start_idx=2)
                if ap and ap > 0:
                    ch = round((jido + edu_exp) / ap * 0.1, 1)
                    slide_and_set(entry, "ch", ch, start_idx=2)
                else:
                    warnings.append(f"{fname}:{name} 年齢階級別人口が見つかりません（ch未更新）")
            else:
                warnings.append(f"{fname}:{name} 目的別歳出が見つかりません（edu/ch未更新）")

        path.write_text(json.dumps(db, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  {fname} 更新完了（{len(db)}件）")

    if warnings:
        print(f"\n⚠️ 警告 {len(warnings)}件（該当自治体は前回値のまま据え置き）:")
        for w in warnings[:30]:
            print("  -", w)
        if len(warnings) > 30:
            print(f"  ...ほか{len(warnings) - 30}件")

    print("\n④ 整合性チェック中...")
    ok = validate()
    if not ok:
        print("\n❌ 検証エラーがあったため処理を中断しました。GitHubへはコミットされません。")
        sys.exit(1)

    print("\n✅ 全ての検証をパスしました。コミットして問題ありません。")


def validate():
    ok = True
    total = 0
    seen_keys = set()
    for fname in REGION_FILES:
        path = ROOT / fname
        try:
            db = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            print(f"  ❌ {fname}: JSON構文エラー {e}")
            ok = False
            continue
        total += len(db)
        for k in db:
            if k in seen_keys:
                print(f"  ❌ 重複キー: {k}")
                ok = False
            seen_keys.add(k)
        for k, e in db.items():
            f, x, d = e.get("f"), e.get("x"), e.get("d")
            if f is not None and not (0 <= f <= 3):
                print(f"  ❌ {fname}:{k} f異常値 {f}")
                ok = False
            if x is not None and not (30 <= x <= 160):
                print(f"  ❌ {fname}:{k} x異常値 {x}")
                ok = False
            if d is not None and not (-20 <= d <= 100):
                print(f"  ❌ {fname}:{k} d異常値 {d}")
                ok = False
            pop = e.get("pop")
            if pop is not None and (not isinstance(pop, (int, float)) or pop <= 0):
                print(f"  ❌ {fname}:{k} pop異常値 {pop}")
                ok = False
            r = e.get("r")
            if r is not None and not (0 <= r <= 50000):
                print(f"  ❌ {fname}:{k} r(財政調整基金)異常値 {r} ※単位換算 miss の可能性")
                ok = False
            eo = e.get("eo")
            if eo is not None and not (0 <= eo <= 200000):
                print(f"  ❌ {fname}:{k} eo(歳出)異常値 {eo} ※単位換算missの可能性")
                ok = False
    print(f"  合計 {total} 件のJSON構文・値域チェック完了")

    # 重複スライド検知：直近の履歴スロットが主値と一致する割合が異常に高い場合、
    # 「同じデータで誤って再実行し履歴が重複した」可能性が高いため警告する
    growth_fields = {"r": ("財政調整基金", 2), "ch": ("子ども1人当たり投資額", 2),
                      "g": ("人口増減率", 2), "edu": ("教育費比率", 2)}
    for fname in REGION_FILES:
        path = ROOT / fname
        db = json.loads(path.read_text(encoding="utf-8"))
        for field, (label, start_idx) in growth_fields.items():
            total_f = 0
            dup_f = 0
            for e in db.values():
                if field not in e:
                    continue
                # 直近の履歴スロット（最も大きい_rN）を動的に探す
                n = start_idx
                newest_key = None
                while f"{field}_r{n}" in e:
                    newest_key = f"{field}_r{n}"
                    n += 1
                if newest_key is None:
                    continue
                total_f += 1
                if e[field] == e[newest_key]:
                    dup_f += 1
            if total_f > 20 and dup_f / total_f > 0.3:
                print(f"  ❌ {fname}: {label}の履歴に重複スライドの疑い（{dup_f}/{total_f}件が主値と完全一致）")
                ok = False
    return ok


if __name__ == "__main__":
    main()
